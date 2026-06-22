"""
Synchronisatie-script: website-data <-> Excel.

De website gebruikt één bronbestand: data/books.js. Elke boektip bevat een
grades-lijst, zodat hetzelfde boek aan meerdere graden gekoppeld kan zijn.

Gebruik:
  uv run --with openpyxl python sync.py export
  uv run --with openpyxl python sync.py import

Bij import worden zowel het nieuwe tabblad "Boeken" als het oude formaat met
drie graadtabbladen ondersteund. Oude dubbele rijen worden daarbij samengevoegd.
"""

import json
import os
import re
import sys

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


ROOT_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(ROOT_DIR, "data")
DATA_FILE = os.path.join(DATA_DIR, "books.js")
EXCEL_FILE = os.path.join(ROOT_DIR, "boekentips_data.xlsx")
DATA_VARIABLE = "BOOKS_DATA"
SHEET_NAME = "Boeken"

GRADES = {
    "graad_1": "1ste Graad",
    "graad_2": "2de Graad",
    "graad_3": "3de Graad",
}

COLUMNS = [
    ("grades", "Graden"),
    ("title", "Titel"),
    ("author", "Auteur"),
    ("isbn", "ISBN"),
    ("pages", "Pagina's"),
    ("publication_date", "Publicatiejaar"),
    ("original_language", "Oorspronkelijke taal"),
    ("genre", "Genre"),
    ("publisher", "Uitgever"),
    ("synopsis", "Synopsis"),
    ("cover_front", "Cover URL (voorkant)"),
    ("cover_back", "Cover URL (achterkant)"),
    ("age_category", "Leeftijdscategorie"),
    ("publisher_url", "Uitgever URL"),
]

GRADE_ALIASES = {
    "1": "graad_1",
    "1ste graad": "graad_1",
    "graad_1": "graad_1",
    "2": "graad_2",
    "2de graad": "graad_2",
    "graad_2": "graad_2",
    "3": "graad_3",
    "3de graad": "graad_3",
    "graad_3": "graad_3",
}


def load_js_data():
    """Lees de JSON-array uit data/books.js."""
    with open(DATA_FILE, "r", encoding="utf-8") as handle:
        content = handle.read()
    return json.loads(content[content.index("["):content.rindex("]") + 1])


def save_js_data(books):
    """Schrijf alle boeken naar het ene JS-databestand."""
    os.makedirs(DATA_DIR, exist_ok=True)
    with open(DATA_FILE, "w", encoding="utf-8") as handle:
        handle.write(f"const {DATA_VARIABLE} = ")
        json.dump(books, handle, ensure_ascii=False, indent=2)
        handle.write(";\n")


def normalise_grades(value, default_grade=None):
    values = value if isinstance(value, list) else re.split(r"[,;]", str(value or ""))
    grades = []
    if default_grade:
        grades.append(default_grade)
    for item in values:
        grade = GRADE_ALIASES.get(str(item).strip().casefold())
        if grade and grade not in grades:
            grades.append(grade)
    return grades


def book_key(book):
    isbn = re.sub(r"[^0-9Xx]", "", str(book.get("isbn") or "")).casefold()
    if isbn:
        return f"isbn:{isbn}"
    title = str(book.get("title") or "").strip().casefold()
    author = str(book.get("author") or "").strip().casefold()
    return f"name:{title}|{author}"


def merge_books(rows):
    """Voeg dubbele boeken samen en verenig hun graadsleutels."""
    merged = []
    by_key = {}
    for incoming in rows:
        key = book_key(incoming)
        if key not in by_key:
            incoming["grades"] = normalise_grades(incoming.get("grades"))
            by_key[key] = incoming
            merged.append(incoming)
            continue
        book = by_key[key]
        book["grades"] = normalise_grades(book["grades"] + incoming.get("grades", []))
        for field, _ in COLUMNS:
            if field != "grades" and not book.get(field) and incoming.get(field):
                book[field] = incoming[field]
    return merged


def export_to_excel():
    """Exporteer data/books.js naar één werkblad."""
    books = load_js_data()
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = SHEET_NAME

    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="8B5E3C", end_color="8B5E3C", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_align = Alignment(vertical="top", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )

    for column_index, (_, label) in enumerate(COLUMNS, 1):
        cell = sheet.cell(row=1, column=column_index, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    for row_index, book in enumerate(books, 2):
        for column_index, (field, _) in enumerate(COLUMNS, 1):
            value = book.get(field)
            if field == "grades":
                value = ", ".join(normalise_grades(value))
            elif value is None:
                value = ""
            cell = sheet.cell(row=row_index, column=column_index, value=value)
            cell.alignment = cell_align
            cell.border = thin_border

    widths = {
        "grades": 20,
        "title": 30,
        "author": 22,
        "isbn": 16,
        "pages": 10,
        "publication_date": 14,
        "original_language": 18,
        "genre": 28,
        "publisher": 22,
        "synopsis": 60,
        "cover_front": 35,
        "cover_back": 35,
        "age_category": 16,
        "publisher_url": 45,
    }
    for column_index, (field, _) in enumerate(COLUMNS, 1):
        sheet.column_dimensions[get_column_letter(column_index)].width = widths.get(field, 15)

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{len(books) + 1}"
    workbook.save(EXCEL_FILE)
    print(f"  {len(books)} unieke boeken geëxporteerd naar tabblad '{SHEET_NAME}'")
    print(f"\nExcel opgeslagen: {EXCEL_FILE}")


def read_sheet(sheet, default_grade=None):
    headers = {}
    for column_index in range(1, sheet.max_column + 1):
        header = sheet.cell(row=1, column=column_index).value
        if not header:
            continue
        for field, label in COLUMNS:
            if label == str(header).strip():
                headers[column_index] = field
                break

    rows = []
    for row_index in range(2, sheet.max_row + 1):
        book = {}
        for column_index, field in headers.items():
            value = sheet.cell(row=row_index, column=column_index).value
            if field == "grades":
                value = normalise_grades(value, default_grade)
            elif field in ("pages", "publication_date"):
                try:
                    value = int(float(str(value))) if value not in (None, "") else None
                except (TypeError, ValueError):
                    value = None
            elif field in ("cover_front", "cover_back", "publisher_url"):
                value = str(value).strip() if value and str(value).strip() else None
            else:
                value = str(value).strip() if value is not None else ""
            book[field] = value

        if default_grade and "grades" not in book:
            book["grades"] = [default_grade]
        if book.get("title"):
            rows.append(book)
    return rows


def import_from_excel():
    """Importeer het nieuwe of oude Excel-formaat naar data/books.js."""
    if not os.path.exists(EXCEL_FILE):
        print(f"FOUT: Excel bestand niet gevonden: {EXCEL_FILE}")
        sys.exit(1)

    workbook = openpyxl.load_workbook(EXCEL_FILE)
    if SHEET_NAME in workbook.sheetnames:
        rows = read_sheet(workbook[SHEET_NAME])
    else:
        rows = []
        for grade, legacy_sheet_name in GRADES.items():
            if legacy_sheet_name in workbook.sheetnames:
                rows.extend(read_sheet(workbook[legacy_sheet_name], grade))

    books = merge_books(rows)
    save_js_data(books)
    print(f"  {len(rows)} rijen ingelezen, {len(books)} unieke boeken opgeslagen")
    print(f"\nJS-bestand bijgewerkt: {DATA_FILE}")


def main():
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(1)
    command = sys.argv[1].lower()
    if command == "export":
        print("Exporteren: JS -> Excel...")
        export_to_excel()
    elif command == "import":
        print("Importeren: Excel -> JS...")
        import_from_excel()
    else:
        print(f"Onbekend commando: {command}")
        print(__doc__)
        sys.exit(1)


if __name__ == "__main__":
    main()
